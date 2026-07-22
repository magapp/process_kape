#!bin/env/bin/python3
"""
build_timeline.py - Merges KAPE-parsed CSV files into a unified forensic timeline.

Reads all CSV files in --process-dir that contain the required KAPE headers
(Timestamp, Event, EventDescription, User, Hostname, ...), normalises timestamps,
optionally enriches each row with IP geolocation data from a MaxMind Enterprise
database, and writes two output files:

  Timeline.csv   — UTF-8 CSV with all events sorted by source file
  Timeline.xlsx  — Excel workbook (sheet "Timeline") with the same data and
                   row background colours based on the Event value

IP enrichment requires ~/.ip-db-full.mmb (MaxMind Enterprise .mmdb). If the
file is absent the Kape_IP / Kape_Country / ... columns are written empty.

Usage:
    python build_timeline.py --process-dir <dir> [options]

Options:
    --process-dir   Directory containing the merged KAPE CSV files (required)
    --csv-output    Output CSV filename, relative to process-dir  (default: Timeline.csv)
    --xlsx-output   Output Excel filename, relative to process-dir (default: Timeline.xlsx)
    --cutoff-days   Drop events older than N days from today       (default: 365)
"""
import argparse
import csv
import ipaddress
import re
from datetime import datetime, timedelta
from pathlib import Path

import chardet
from dateutil import parser as dateutil_parser
from geoip2 import database
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

IP_DB_FILE = Path.home() / ".ip-db-full.mmb"

IPV4_RE = re.compile(r"\b(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})\b")
IPV6_RE = re.compile(r"\b(?:[0-9a-fA-F]{1,4}:){2,7}[0-9a-fA-F]{1,4}\b")

REQUIRED_HEADERS = ["Timestamp", "Event", "EventDescription", "User", "Hostname",
                    "IPAddress", "RemoteHostname", "RemoteIP", "Source", "Comment"]

# (substring, ARGB hex color, swedish description) — first match wins, case-insensitive
EVENT_COLORS = [
    ("eventlog cleared",     "FFADD8E6", "Händelseloggen har rensats – kan tyda på att ett spår försökt döljas"),
    ("createremotethread",   "FFFF6666", "En process har skapat en tråd i en annan process – vanlig injektionsteknik"),
    ("antivirus",            "FFFF4444", "Antiviruset har detekterat eller blockerat ett hot"),
    ("failed logon",         "FF6080CC", "Misslyckat inloggningsförsök – kan indikera brute force eller felaktiga credentials"),
    ("successful logon",     "FF87CEEB", "Lyckad inloggning på systemet"),
    ("incoming network",     "FFFFA500", "Inkommande nätverksanslutning upprättad till denna dator"),
    ("outgoing network",     "FFD2B48C", "Utgående nätverksanslutning från denna dator"),
    ("smb server",           "FFE0EFFF", "Inkommande SMB-anslutning – fildelning eller potentiell lateral rörelse"),
    ("smb client",           "FFB3CEED", "Utgående SMB-anslutning – åtkomst till fjärrresurs"),
    ("network connection",   "FFCCE5FF", "Nätverksanslutning (riktning okänd)"),
    ("rdp server",           "FFFFCC66", "Någon har anslutit till denna dator via RDP"),
    ("rdp client",           "FFFFDD99", "Denna dator har anslutit till en annan dator via RDP"),
    ("account creation",     "FFFFFF99", "Ett nytt användarkonto har skapats"),
    ("service installation", "FFCC99FF", "En ny Windows-tjänst har installerats"),
    ("scheduled task",       "FFDD99FF", "En schemalagd uppgift har skapats eller ändrats"),
    ("persistance",          "FFEEAAFF", "Persistens-mekanism identifierad – program som startar automatiskt"),
    ("powershell",           "FF5B8CB8", "PowerShell har körts – kan indikera skript eller automation"),
    ("process creation",     "FFB3FFB3", "En ny process har startats"),
    ("amcache",              "FFCCFFCC", "Program som körts, spårat via Amcache"),
    ("appcompatcache",       "FFCCEECC", "Program som körts, spårat via AppCompatCache (ShimCache)"),
    ("email",                "FFD4A96A", "E-postrelaterad aktivitet"),
    ("filecreated",          "FFAADDFF", "En fil har skapats på systemet"),
    ("filemodified",         "FFE8E8E8", "En fil har modifierats på systemet"),
    ("link file",            "FFFFFF99", "En LNK-fil (genväg) har använts – indikerar filåtkomst"),
    ("shellbags",            "FFFFD966", "En mapp har öppnats i Utforskaren, spårat via Shellbags"),
    ("registrylastwrite",    "FFCCCCCC", "En registernyckel har senast skrivits"),
    ("registry",             "FFDDDDDD", "En registernyckel har skrivits eller lästs"),
    ("browsing history",     "FFAAFFCC", "Webbsida besökt i en webbläsare"),
]


def get_row_fill(event_value):
    """Return a PatternFill for the given Event value, or None if no rule matches.

    Iterates EVENT_COLORS in order; the first substring that is found
    (case-insensitive) in event_value wins.
    """
    v = event_value.lower()
    for substring, color, _description in EVENT_COLORS:
        if substring in v:
            return PatternFill(fill_type="solid", fgColor=color)
    return None


def has_required_headers(header):
    """Return True if the CSV header starts with all REQUIRED_HEADERS in order."""
    return header[:len(REQUIRED_HEADERS)] == REQUIRED_HEADERS


# ---------------------------------------------------------------------------
# Raw KAPE file adapters
# Each adapter maps raw EZTools columns to REQUIRED_HEADERS order.
# ---------------------------------------------------------------------------

def _colmap(header):
    """Return a dict mapping column name (BOM-stripped) → index."""
    return {name.lstrip("\ufeff"): i for i, name in enumerate(header)}


def _g(row, cm, col, default=""):
    """Get a cell value by column name; return default if missing."""
    i = cm.get(col, -1)
    return (row[i] if 0 <= i < len(row) else default) or default


def _detect(required_cols):
    """Return a detector function that checks required_cols are all present."""
    def detect(header):
        stripped = {c.lstrip("\ufeff") for c in header}
        return all(c in stripped for c in required_cols)
    return detect


def _adapt_evtx(header, row):
    """EvtxECmd_Output.csv → timeline row."""
    cm = _colmap(header)
    map_desc = _g(row, cm, "MapDescription")
    event_id  = _g(row, cm, "EventId")
    event = map_desc if map_desc else f"EventID {event_id}"
    p1 = _g(row, cm, "PayloadData1")
    p2 = _g(row, cm, "PayloadData2")
    desc = " | ".join(p for p in [map_desc, p1, p2] if p) or event
    remote_host = _g(row, cm, "RemoteHost")
    remote_ip   = extract_ip(remote_host)
    remote_hn   = remote_host if remote_host and not remote_ip else ""
    return [
        _g(row, cm, "TimeCreated"),
        event,
        desc,
        _g(row, cm, "UserName"),
        _g(row, cm, "Computer"),
        "",
        remote_hn,
        remote_ip,
        _g(row, cm, "Channel"),
        p1,
    ]


def _adapt_mft(header, row):
    """MFTECmd_$MFT_Output.csv → timeline row (files only, skip dirs)."""
    cm = _colmap(header)
    if _g(row, cm, "IsDirectory", "False").lower() in ("true", "1"):
        return None  # skip directories
    if _g(row, cm, "InUse", "True").lower() not in ("true", "1"):
        return None  # skip deleted entries
    parent = _g(row, cm, "ParentPath")
    fname  = _g(row, cm, "FileName")
    path   = f"{parent}\\{fname}" if parent else fname
    return [
        _g(row, cm, "LastModified0x10") or _g(row, cm, "Created0x10"),
        "File System",
        path,
        "",
        "",
        "",
        "",
        "",
        "MFT",
        "",
    ]


def _adapt_usn(header, row):
    """MFTECmd_$J_Output.csv (USN Journal) → timeline row."""
    cm = _colmap(header)
    parent = _g(row, cm, "ParentPath")
    fname  = _g(row, cm, "Name")
    path   = f"{parent}\\{fname}" if parent else fname
    reasons = _g(row, cm, "UpdateReasons")
    return [
        _g(row, cm, "UpdateTimestamp"),
        "USN Journal",
        path,
        "",
        "",
        "",
        "",
        "",
        "USN Journal",
        reasons,
    ]


def _adapt_leCmd(header, row):
    """LECmd_Output.csv (LNK files) → timeline row."""
    cm   = _colmap(header)
    path = _g(row, cm, "LocalPath") or _g(row, cm, "TargetIDAbsolutePath")
    ts   = _g(row, cm, "TargetModified") or _g(row, cm, "SourceModified")
    src  = _g(row, cm, "SourceFile")
    # Extract username from SourceFile path (e.g. C:\Users\USERNAME\...)
    user = ""
    m = re.search(r"[Uu]sers[/\\]([^/\\]+)", src)
    if m:
        user = m.group(1)
    return [
        ts,
        "Link File",
        path,
        user,
        _g(row, cm, "MachineID"),
        "",
        "",
        "",
        "LNK",
        "",
    ]


def _adapt_autodest(header, row):
    """AutomaticDestinations.csv (JumpLists) → timeline row."""
    cm   = _colmap(header)
    path = _g(row, cm, "Path") or _g(row, cm, "LocalPath") or _g(row, cm, "TargetIDAbsolutePath")
    ts   = _g(row, cm, "LastModified") or _g(row, cm, "TargetModified")
    src  = _g(row, cm, "SourceFile")
    user = ""
    m = re.search(r"[Uu]sers[/\\]([^/\\]+)", src)
    if m:
        user = m.group(1)
    return [
        ts,
        "Link File",
        path,
        user,
        _g(row, cm, "Hostname"),
        "",
        "",
        "",
        "AutomaticDestinations",
        _g(row, cm, "AppIdDescription"),
    ]


def _adapt_appcompat(header, row):
    """AppCompatCache (ShimCache) → timeline row."""
    cm = _colmap(header)
    return [
        _g(row, cm, "LastModifiedTimeUTC"),
        "AppCompatCache",
        _g(row, cm, "Path"),
        "",
        "",
        "",
        "",
        "",
        "AppCompatCache",
        _g(row, cm, "Executed"),
    ]


def _adapt_browsing(header, row):
    """BrowsingHistory.csv → timeline row."""
    cm    = _colmap(header)
    url   = _g(row, cm, "URL")
    title = _g(row, cm, "Title")
    desc  = f"{url} ({title})" if title else url
    return [
        _g(row, cm, "Visit Time"),
        "Browsing History",
        desc,
        _g(row, cm, "User Profile"),
        "",
        "",
        "",
        "",
        _g(row, cm, "Web Browser"),
        "",
    ]


# Ordered list of (detector, adapt_fn) — first match wins
RAW_ADAPTERS = [
    (_detect({"TimeCreated", "EventId", "Channel", "MapDescription"}), _adapt_evtx),
    (_detect({"EntryNumber", "FileName", "Created0x10", "IsDirectory"}),  _adapt_mft),
    (_detect({"UpdateTimestamp", "UpdateReasons", "EntryNumber"}),         _adapt_usn),
    (_detect({"TargetIDAbsolutePath", "MachineID", "TrackerCreatedOn"}),  _adapt_leCmd),
    (_detect({"AppId", "AppIdDescription", "LastUsedEntryNumber"}),       _adapt_autodest),
    (_detect({"CacheEntryPosition", "LastModifiedTimeUTC", "Executed"}),  _adapt_appcompat),
    (_detect({"URL", "Visit Time", "Web Browser"}),                        _adapt_browsing),
]


def get_raw_adapter(header):
    """Return the first matching raw adapter function, or None."""
    for detect, adapt in RAW_ADAPTERS:
        if detect(header):
            return adapt
    return None


def is_public_ip(ip_str):
    """Return True if ip_str is a globally routable (public) IP address."""
    try:
        if ip_str.split(".")[3] == "0":
            return False
    except Exception:
        pass
    try:
        return ipaddress.ip_address(ip_str).is_global
    except ValueError:
        return False


def looks_like_filepath(value):
    """Return True if value contains a path separator (/ or \\)."""
    return bool(re.search(r'[/\\]', value))


def ip_is_embedded(value, match):
    """Return True if the IP match is embedded inside a larger token."""
    start, end = match.start(), match.end()
    before = value[start - 1] if start > 0 else " "
    after = value[end] if end < len(value) else " "
    if before == "." or after == ".":
        return True
    clean_boundary = {" ", ",", "\t", '"', "'", "(", ")", "[", "]", "", ":", ">", "<"}
    return before not in clean_boundary and after not in clean_boundary


def _looks_like_version(ip_str, cell_value):
    """Return True if the dotted-quad looks like a version number rather than a routable IP."""
    octets = ip_str.split(".")
    if any(o == "0" for o in octets[:3]):
        return True
    if cell_value.strip() == ip_str and int(octets[0]) <= 9 and int(octets[1]) <= 9:
        return True
    return False


def extract_ip(value):
    """Extract the first public IPv4 or IPv6 address from value, or return ''."""
    if looks_like_filepath(value) or "version" in value.lower() or "revision" in value.lower() or "oid" in value.lower():
        return ""
    for m in IPV4_RE.finditer(value):
        ip = m.group(0)
        if not all(0 <= int(g) <= 255 for g in m.groups()):
            continue
        if not is_public_ip(ip):
            continue
        if ip_is_embedded(value, m):
            continue
        if _looks_like_version(ip, value):
            continue
        return ip
    for m in IPV6_RE.finditer(value):
        if is_public_ip(m.group(0)):
            return m.group(0)
    return ""


def lookup_ip(db, ip):
    """Look up ip in the MaxMind Enterprise database.

    Returns (country, city, connection_type, isp) as strings.
    Returns four empty strings if the lookup fails or ip is empty.
    """
    if not ip:
        return "", "", "", ""
    try:
        res = db.enterprise(ip)
        country = str(res.country.names.get("en", "")).replace("None", "")
        city    = str(res.city.names.get("en", "")).replace("None", "")
        ip_type = str(res.traits.connection_type).replace("None", "")
        isp     = str(res.traits.isp).replace("None", "")
        return country, city, ip_type, isp
    except Exception:
        return "", "", "", ""


_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize(value):
    """Strip illegal Excel control characters from string values."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub("", value)
    return value


def write_xlsx(path, header, rows, ip_data=None):
    """Write rows to an Excel workbook at path.

    Creates a single sheet named "Timeline" with a bold header row.
    Each data row is colour-coded via get_row_fill() based on its Event value.
    Column widths are auto-fitted (capped at 60 characters).

    Args:
        path:    Destination file path.
        header:  List of column names.
        rows:    List of (row_data, event_value) tuples.
        ip_data: Dict {ip: [country, city, ip_type, isp, count]} for the IP sheet.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Timeline")

    col_widths = [min(len(str(val)), 60) for val in header]

    header_font = Font(bold=True)
    header_cells = []
    for val in header:
        cell = WriteOnlyCell(ws, value=val)
        cell.font = header_font
        header_cells.append(cell)
    ws.append(header_cells)

    for row_data, event_value in rows:
        for i, val in enumerate(row_data):
            if i < len(col_widths):
                col_widths[i] = min(max(col_widths[i], len(str(val or ""))), 60)
        fill = get_row_fill(event_value)
        if fill:
            out_cells = []
            for val in row_data:
                cell = WriteOnlyCell(ws, value=_sanitize(val))
                cell.fill = fill
                out_cells.append(cell)
            ws.append(out_cells)
        else:
            ws.append([_sanitize(v) for v in row_data])

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    # --- Färger sheet ---
    wc = wb.create_sheet("Färger och förklaring")
    col_a = max(len("Event"),       max(len(s) for s, _, __ in EVENT_COLORS))
    col_b = max(len("Beskrivning"), max(len(d) for _, __, d in EVENT_COLORS))
    wc.column_dimensions["A"].width = col_a + 2
    wc.column_dimensions["B"].width = min(col_b + 2, 100)

    bold = Font(bold=True)
    h1 = WriteOnlyCell(wc, value="Event");       h1.font = bold
    h2 = WriteOnlyCell(wc, value="Beskrivning"); h2.font = bold
    wc.append([h1, h2])

    for substring, color, description in EVENT_COLORS:
        fill = PatternFill(fill_type="solid", fgColor=color)
        c_event = WriteOnlyCell(wc, value=substring)
        c_event.fill = fill
        c_desc  = WriteOnlyCell(wc, value=description)
        wc.append([c_event, c_desc])

    wc.append([])
    wc.append(["I fliken IP-adresser finns en lista på publika IP-adresser som har hittats."])

    # --- IP-adresser sheet ---
    if ip_data:
        wi = wb.create_sheet("IP-adresser")
        ip_headers = ["IP address", "Nbr", "Country", "City", "Type", "ISP"]
        wi.column_dimensions["A"].width = 18
        wi.column_dimensions["B"].width = 6
        wi.column_dimensions["C"].width = 20
        wi.column_dimensions["D"].width = 20
        wi.column_dimensions["E"].width = 18
        wi.column_dimensions["F"].width = 30


        bold = Font(bold=True)
        header_cells = []
        for val in ip_headers:
            cell = WriteOnlyCell(wi, value=val)
            cell.font = bold
            header_cells.append(cell)
        wi.append(header_cells)

        for ip, (country, city, ip_type, isp, count) in sorted(ip_data.items()):
            wi.append([ip, count, country, city, ip_type, isp])

    wb.save(path)


def parse_timestamp(value):
    """Parse a timestamp string, returning a datetime or None.

    Tries datetime.fromisoformat first (fast), then falls back to dateutil.
    Catches ValueError, TypeError and OverflowError so no input can hang or crash.
    """
    try:
        return datetime.fromisoformat(value)
    except (ValueError, TypeError):
        pass
    try:
        return dateutil_parser.parse(value)
    except (ValueError, TypeError, OverflowError):
        return None


def main():
    parser = argparse.ArgumentParser(description="Create a unified timeline from KAPE CSV files")
    parser.add_argument("--process-dir",  required=True,          help="Directory with CSV files")
    parser.add_argument("--csv-output",   default="Timeline.csv",  help="Output CSV filename (default: Timeline.csv)")
    parser.add_argument("--xlsx-output",  default="Timeline.xlsx", help="Output Excel filename (default: Timeline.xlsx)")
    parser.add_argument("--cutoff-days",  default=365, type=int,   help="Ignore events older than N days (default: 365)")
    args = parser.parse_args()

    process_dir = Path(args.process_dir)
    csv_path    = process_dir / args.csv_output
    xlsx_path   = process_dir / args.xlsx_output
    cutoff      = datetime.now() - timedelta(days=args.cutoff_days)

    # Prevent the CSV parser from hanging on malformed fields (e.g. unclosed quotes
    # in service descriptions). Raises csv.Error if a field exceeds this size.
    csv.field_size_limit(10 * 1024 * 1024)  # 10 MB

    if not IP_DB_FILE.exists():
        print(f"Warning: IP database not found at {IP_DB_FILE}, skipping IP enrichment")
        db_ctx = None
    else:
        try:
            db_ctx = database.Reader(IP_DB_FILE)
        except Exception as e:
            print(f"Warning: Could not open IP database ({e}), skipping IP enrichment")
            db_ctx = None

    skip_files = {args.csv_output, args.xlsx_output}

    xlsx_header = None
    xlsx_rows   = []
    ip_data     = {}   # ip -> [country, city, ip_type, isp, count]

    for csv_file in sorted(process_dir.glob("*.csv")):
        if csv_file.name in skip_files:
            continue

        try:
            detected = chardet.detect(csv_file.read_bytes())["encoding"]
            encoding = detected if detected and detected.lower() not in ("ascii", "utf-8") else "utf-8-sig"
            with open(csv_file, newline="", encoding=encoding, errors="replace") as f:
                reader = csv.reader(f)
                try:
                    header = next(reader)
                except StopIteration:
                    continue

                adapter = None
                if not has_required_headers(header):
                    adapter = get_raw_adapter(header)
                    if adapter is None:
                        print(f"Skipping {csv_file.name} (missing required headers)")
                        continue

                print(f"Adding {csv_file.name} {'(raw adapter)' if adapter else ''}...")

                if xlsx_header is None:
                    xlsx_header = REQUIRED_HEADERS + ["Kape_IP", "Kape_Country", "Kape_City", "Kape_Type", "Kape_ISP"]

                for row in reader:
                    if not row:
                        continue

                    if adapter:
                        mapped = adapter(header, row)
                        if mapped is None:
                            continue
                        row = mapped
                    else:
                        row = list(row)

                    ts = parse_timestamp(row[0])
                    if ts is None:
                        continue
                    row[0] = ts.strftime("%Y-%m-%d %H:%M:%S")
                    if ts.replace(tzinfo=None) < cutoff:
                        continue

                    ip = next((v for cell in row if (v := extract_ip(cell))), "") if db_ctx else ""
                    country, city, ip_type, isp = lookup_ip(db_ctx, ip) if db_ctx else ("", "", "", "")
                    if ip:
                        if ip not in ip_data:
                            ip_data[ip] = [country, city, ip_type, isp, 0]
                        ip_data[ip][4] += 1

                    full_row    = row + [ip, country, city, ip_type, isp]
                    event_value = row[1] if len(row) > 1 else ""
                    xlsx_rows.append((full_row, event_value))

        except csv.Error as e:
            print(f"Warning: skipping {csv_file.name} due to CSV parse error: {e}")

    if db_ctx:
        db_ctx.close()

    # Prefer the comprehensive IP list produced by build_ip_lists.py (which scans all raw
    # CSV files, not just timeline rows) if it exists in the process directory.
    ip_addresses_csv = process_dir / "ip-addresses.csv"
    if ip_addresses_csv.exists():
        ip_data = {}
        with open(ip_addresses_csv, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                ip = row.get("IPAddress", "").strip()
                if ip:
                    ip_data[ip] = [
                        row.get("Country", ""),
                        row.get("City", ""),
                        row.get("Type", ""),
                        row.get("ISP", ""),
                        int(row.get("Nbr occurance", 0) or 0),
                    ]
        print(f"Loaded {len(ip_data)} IPs from {ip_addresses_csv.name} for Excel IP tab")

    if xlsx_header:
        xlsx_rows.sort(key=lambda item: item[0][0])

        with open(csv_path, "w", newline="", encoding="utf-8-sig") as out:
            writer = csv.writer(out)
            writer.writerow(xlsx_header)
            writer.writerows(row for row, _ in xlsx_rows)
        print(f"Wrote {csv_path.name} ({len(xlsx_rows)} rows)")

        write_xlsx(xlsx_path, xlsx_header, xlsx_rows, ip_data=ip_data if ip_data else None)
        print(f"Wrote {xlsx_path.name} ({len(xlsx_rows)} rows, {len(ip_data)} unique IPs)")


if __name__ == "__main__":
    main()
